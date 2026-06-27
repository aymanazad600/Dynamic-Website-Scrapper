"""Hybrid Kaggle Dataset Selector.

High-quality workflow:
1. Open the Kaggle search page with Playwright and read the visible result cards
   the same way you see them in the browser.
2. Keep Kaggle dataset results by default.
3. Use Kaggle API/downloaded files to enrich each row with real file names,
   file types, row counts, and column headers.
4. Export a polished Excel workbook.

This script uses public browser-visible content plus Kaggle's official API.
It does not bypass captcha, login walls, paywalls, or anti-bot protection.

Input:
    urls.txt can contain one Kaggle search URL, for example:
    https://www.kaggle.com/search?q=Bank+Customer+Churn+Dataset

Output:
    output/kaggle_dataset_selector.xlsx
"""

from __future__ import annotations

import argparse
import os
import re
from pathlib import Path
from typing import Any
from urllib.parse import parse_qs, quote_plus, urlsplit

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


BASE_DIR = Path(__file__).resolve().parent
URLS_FILE = BASE_DIR / "urls.txt"
OUTPUT_FILE = BASE_DIR / "output" / "kaggle_dataset_selector.xlsx"
DOWNLOAD_DIR = BASE_DIR / "work" / "kaggle_downloads"

MAX_DATASETS = 20
MAX_HEADER_FILES_PER_DATASET = 8
MAX_DATASET_BYTES = 300 * 1024 * 1024
KAGGLE_BASE_URL = "https://www.kaggle.com"

COLUMNS = [
    "search_rank",
    "dataset_title",
    "dataset_ref",
    "result_url",
    "result_type",
    "visible_author",
    "visible_age",
    "visible_votes",
    "visible_downloads",
    "visible_comments",
    "visible_summary",
    "file_name",
    "file_types",
    "size_mb",
    "api_author",
    "updated_date",
    "api_votes",
    "api_downloads",
    "api_views",
    "usability",
    "license",
    "update_frequency",
    "file_count",
    "column_count",
    "row_count_sample",
    "column_headers",
    "target_column_guess",
    "about_dataset",
    "tags",
    "business_problem",
    "project_fit_score",
    "recommendation",
    "metadata_notes",
]

BUSINESS_WORDS = {
    "bank", "banking", "customer", "churn", "loan", "credit", "fraud", "risk",
    "finance", "financial", "transaction", "transactions", "marketing", "sales",
    "business", "insurance", "retail", "revenue", "profit",
}
SUPPORTED_DATA_FILES = {".csv", ".tsv", ".xlsx", ".xls", ".json", ".jsonl", ".parquet"}
COOKIE_BUTTON_PATTERN = re.compile(r"^\s*(OK,\s*Got it|Accept|Accept all)\s*$", re.IGNORECASE)


def configure_kaggle_token() -> None:
    """Load Kaggle's newer ~/.kaggle/access_token before importing the Kaggle client."""
    access_token_path = Path.home() / ".kaggle" / "access_token"
    if access_token_path.exists():
        token = access_token_path.read_text(encoding="utf-8").strip()
        if token:
            os.environ.setdefault("KAGGLE_API_TOKEN", token)
            os.environ.setdefault("KAGGLE_USERNAME", "__token__")
            os.environ.setdefault("KAGGLE_KEY", token)


configure_kaggle_token()


def log(message: str) -> None:
    print(message, flush=True)


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def clean_header(value: Any) -> str:
    text = clean_text(value)
    return re.sub(r"^Unnamed:\s*\d+$", "", text, flags=re.IGNORECASE).strip()


def unique_join(values: list[Any], separator: str = ", ", limit: int = 250) -> str:
    seen: set[str] = set()
    output: list[str] = []
    for value in values:
        cleaned = clean_header(value)
        key = cleaned.casefold()
        if cleaned and key not in seen:
            seen.add(key)
            output.append(cleaned)
        if len(output) >= limit:
            break
    return separator.join(output)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Hybrid Kaggle Dataset Selector")
    parser.add_argument("--url", help="Kaggle search URL")
    parser.add_argument("--query", help="Plain Kaggle search phrase")
    parser.add_argument("--max", type=int, default=MAX_DATASETS, help=f"Maximum dataset rows. Default: {MAX_DATASETS}")
    parser.add_argument("--include-notebooks", action="store_true", help="Also include visible /code/ result cards")
    parser.add_argument("--no-download", action="store_true", help="Skip dataset downloads and only use visible/API metadata")
    return parser.parse_args()


def search_text_from_value(value: str) -> str:
    parsed = urlsplit(value)
    query = parse_qs(parsed.query).get("q", [""])[0].strip()
    return query or value.strip()


def read_search_input(args: argparse.Namespace) -> tuple[str, str]:
    """Return search_url and search_text from CLI args or urls.txt."""
    if args.query:
        search_text = args.query.strip()
        return f"{KAGGLE_BASE_URL}/search?q={quote_plus(search_text)}", search_text
    if args.url:
        return args.url.strip(), search_text_from_value(args.url)

    if not URLS_FILE.exists():
        raise FileNotFoundError(f"Missing input file: {URLS_FILE}")
    lines = [
        line.strip()
        for line in URLS_FILE.read_text(encoding="utf-8").splitlines()
        if line.strip() and not line.lstrip().startswith("#")
    ]
    if not lines:
        raise ValueError("urls.txt must contain one Kaggle search URL or search phrase.")

    value = lines[0]
    if value.startswith("http"):
        return value, search_text_from_value(value)
    return f"{KAGGLE_BASE_URL}/search?q={quote_plus(value)}", value


def dataset_search_url(search_text: str) -> str:
    """Kaggle's dataset listing page returns many dataset-only results."""
    return f"{KAGGLE_BASE_URL}/datasets?search={quote_plus(search_text)}"


def result_type_from_url(url: str) -> str:
    path = urlsplit(url).path.lower()
    if path.startswith("/datasets/"):
        return "Dataset"
    if path.startswith("/code/"):
        return "Notebook"
    if path.startswith(("/discussion/", "/discussions/")):
        return "Discussion"
    if path.startswith("/competitions/"):
        return "Competition"
    return ""


def ref_from_url(url: str) -> str:
    parts = [part for part in urlsplit(url).path.split("/") if part]
    if len(parts) >= 3 and parts[0] in {"datasets", "code"}:
        return f"{parts[1]}/{parts[2]}"
    return ""


def kaggle_absolute_url(href: str) -> str:
    if href.startswith("http"):
        return href.split("?", 1)[0].rstrip("/")
    return f"{KAGGLE_BASE_URL}{href}".split("?", 1)[0].rstrip("/")


def parse_metric(text: str, labels: tuple[str, ...]) -> str:
    label_pattern = "|".join(re.escape(label) for label in labels)
    patterns = [
        rf"\b([\d,.]+\s*[KMBkmb]?)\s+(?:{label_pattern})\b",
        rf"\b(?:{label_pattern})\s*[:\-]?\s*([\d,.]+\s*[KMBkmb]?)\b",
    ]
    for pattern in patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            return clean_text(match.group(1))
    return ""


def parse_usability(text: str) -> str:
    match = re.search(r"\bUsability\s*([\d.]+)", text, re.IGNORECASE)
    return clean_text(match.group(1)) if match else ""


def parse_file_count(text: str) -> str:
    match = re.search(r"\b(\d+)\s+Files?\b", text, re.IGNORECASE)
    return clean_text(match.group(1)) if match else ""


def parse_file_types(text: str) -> str:
    values = re.findall(r"\bFiles?\s*\(([^)]+)\)", text, re.IGNORECASE)
    return unique_join(values, separator=", ", limit=10)


def parse_size_mb(text: str) -> str:
    match = re.search(r"\b([\d.]+)\s*(KB|MB|GB)\b", text, re.IGNORECASE)
    if not match:
        return ""
    value = float(match.group(1))
    unit = match.group(2).upper()
    if unit == "KB":
        value = value / 1024
    elif unit == "GB":
        value = value * 1024
    return f"{value:.1f}"


def parse_vote_from_lines(lines: list[str]) -> str:
    """Kaggle result cards show votes as arrow_drop_up followed by a number."""
    for index, line in enumerate(lines):
        if line == "arrow_drop_up" and index + 1 < len(lines):
            candidate = clean_text(lines[index + 1])
            if re.fullmatch(r"[\d,.]+\s*[KMBkmb]?", candidate):
                return candidate
    return ""


def parse_title_from_lines(lines: list[str], anchor_text: str) -> str:
    """Extract the card title while ignoring Material icons like table_chart/code."""
    icon_words = {"table_chart", "code", "comment", "reply", "tenancy", "emoji_events", "arrow_drop_up"}
    for line in lines:
        key = line.casefold()
        if key in icon_words:
            continue
        if re.search(r"^(Dataset|Notebook|Discussion|Competition)\s*·", line, re.IGNORECASE):
            continue
        if re.fullmatch(r"[\d,.]+\s*[KMBkmb]?", line):
            continue
        if len(line) >= 3:
            return line
    return clean_text(anchor_text).splitlines()[0] if anchor_text else ""


def parse_visible_author(text: str) -> str:
    patterns = [
        r"\bby\s+([A-Z][A-Za-z0-9_. -]{1,80})(?:\n|$)",
        r"Dataset\s*·\s*[^·]+·\s*by\s+([^\n]+)",
        r"Notebook\s*·\s*[^·]+·\s*by\s+([^\n]+)",
        r"(?:^|\n)([^\n·]+)\s*·\s*Updated\s+",
    ]
    for pattern in patterns:
        match = re.search(pattern, text)
        if match:
            return clean_text(match.group(1))
    return ""


def parse_visible_age(text: str) -> str:
    match = re.search(r"\b(\d+\s*(?:y|yr|year|years|mo|month|months|d|day|days)\s+ago)\b", text, re.IGNORECASE)
    return clean_text(match.group(1)) if match else ""


def parse_visible_summary(lines: list[str], title: str) -> str:
    skipped = {
        title.casefold(), "dataset", "notebook", "discussion", "competition", "table_chart",
        "code", "comment", "reply", "arrow_drop_up",
    }
    useful: list[str] = []
    for line in lines:
        key = line.casefold()
        if key in skipped:
            continue
        if re.search(r"^(dataset|notebook|discussion)\s*·", line, re.IGNORECASE):
            continue
        if re.search(r"\b(votes?|downloads?|comments?)\b", line, re.IGNORECASE):
            continue
        if len(line) >= 20:
            useful.append(line)
        if len(" ".join(useful)) > 400:
            break
    return " ".join(useful)[:500]


def parse_card(raw: dict[str, Any], rank: int) -> dict[str, Any]:
    url = kaggle_absolute_url(clean_text(raw.get("href")))
    text = clean_text(raw.get("text"))
    lines = [clean_text(line) for line in str(raw.get("text", "")).splitlines() if clean_text(line)]
    title = parse_title_from_lines(lines, clean_text(raw.get("anchorText")))
    result_type = result_type_from_url(url)
    if not result_type:
        result_type = first_line_type(lines)

    return {
        "search_rank": rank,
        "dataset_title": title,
        "dataset_ref": ref_from_url(url),
        "result_url": url,
        "result_type": result_type,
        "visible_author": parse_visible_author("\n".join(lines)),
        "visible_age": parse_visible_age("\n".join(lines)),
        "visible_votes": parse_vote_from_lines(lines) or parse_metric(text, ("vote", "votes")),
        "visible_downloads": parse_metric(text, ("download", "downloads")),
        "visible_comments": parse_metric(text, ("comment", "comments")),
        "visible_summary": parse_visible_summary(lines, title),
        "_visible_usability": parse_usability(text),
        "_visible_file_count": parse_file_count(text),
        "_visible_file_types": parse_file_types(text),
        "_visible_size_mb": parse_size_mb(text),
    }


def first_line_type(lines: list[str]) -> str:
    for line in lines[:4]:
        if "dataset" in line.casefold():
            return "Dataset"
        if "notebook" in line.casefold():
            return "Notebook"
        if "discussion" in line.casefold():
            return "Discussion"
    return ""


def get_attr(obj: Any, names: list[str], default: Any = "") -> Any:
    for name in names:
        if isinstance(obj, dict) and name in obj:
            return obj.get(name, default)
        if hasattr(obj, name):
            value = getattr(obj, name)
            if value is not None:
                return value
    return default


def first_attr(objects: list[Any], names: list[str], default: Any = "") -> Any:
    for obj in objects:
        value = get_attr(obj, names, "")
        if value not in ("", None, []):
            return value
    return default


def not_available(value: Any, fallback: str = "Not provided by Kaggle API") -> str:
    cleaned = clean_text(value)
    return cleaned if cleaned else fallback


def safe_number(value: Any) -> str:
    if value is None or value == "":
        return ""
    return str(value)


def tag_names(tags: Any) -> str:
    if not tags:
        return ""
    if not isinstance(tags, list):
        tags = [tags]
    values: list[str] = []
    for tag in tags:
        values.append(clean_text(get_attr(tag, ["name", "fullPath", "ref"], tag)))
    return unique_join(values, separator=", ", limit=30)


def format_size_mb(size_bytes: int) -> str:
    return f"{size_bytes / (1024 * 1024):.1f}" if size_bytes else ""


def safe_size_bytes(dataset: Any) -> int:
    value = get_attr(dataset, ["totalBytes", "total_bytes", "size"])
    try:
        return int(value)
    except Exception:
        return 0


def data_files(folder: Path) -> list[Path]:
    files = [
        path
        for path in folder.rglob("*")
        if path.is_file() and path.suffix.lower() in SUPPORTED_DATA_FILES
    ]
    return sorted(files, key=lambda path: path.stat().st_size if path.exists() else 0)


def dataset_file_types(files: list[Path]) -> str:
    return unique_join([path.suffix.lower().lstrip(".") for path in files], separator=", ", limit=20)


def headers_from_file(path: Path) -> list[str]:
    suffix = path.suffix.lower()
    try:
        if suffix == ".csv":
            return list(pd.read_csv(path, nrows=0, low_memory=False).columns)
        if suffix == ".tsv":
            return list(pd.read_csv(path, nrows=0, sep="\t", low_memory=False).columns)
        if suffix in {".xlsx", ".xls"}:
            return list(pd.read_excel(path, nrows=0).columns)
        if suffix == ".parquet":
            return list(pd.read_parquet(path).columns)
        if suffix in {".json", ".jsonl"}:
            frame = pd.read_json(path, lines=suffix == ".jsonl", nrows=5)
            return list(frame.columns)
    except Exception:
        return []
    return []


def sample_row_count(path: Path) -> str:
    suffix = path.suffix.lower()
    try:
        if suffix in {".csv", ".tsv"}:
            with path.open("r", encoding="utf-8", errors="ignore") as handle:
                return str(max(0, sum(1 for _ in handle) - 1))
        if suffix in {".xlsx", ".xls"}:
            return str(len(pd.read_excel(path, usecols=[0])))
        if suffix == ".jsonl":
            with path.open("r", encoding="utf-8", errors="ignore") as handle:
                return str(sum(1 for _ in handle))
        if suffix == ".json":
            return str(len(pd.read_json(path)))
        if suffix == ".parquet":
            return str(len(pd.read_parquet(path, columns=[])))
    except Exception:
        return ""
    return ""


def guess_target_column(headers: str) -> str:
    candidates = [
        "churn", "exited", "target", "label", "class", "y", "outcome",
        "default", "fraud", "is_fraud", "approved", "response", "subscribed",
    ]
    header_values = [value.strip() for value in headers.split(",") if value.strip()]
    lower_map = {value.casefold(): value for value in header_values}
    for candidate in candidates:
        if candidate in lower_map:
            return lower_map[candidate]
    for value in header_values:
        lowered = value.casefold()
        if any(candidate in lowered for candidate in candidates if len(candidate) > 1):
            return value
    return ""


def extract_headers(files: list[Path]) -> tuple[str, str, str, str, str, str]:
    selected_files = files[:MAX_HEADER_FILES_PER_DATASET]
    file_names = [path.name for path in selected_files]
    headers: list[str] = []
    for path in selected_files:
        headers.extend(headers_from_file(path))

    header_text = unique_join(headers)
    column_count = str(len([value for value in header_text.split(", ") if value.strip()])) if header_text else ""
    row_count = ""
    for path in selected_files:
        row_count = sample_row_count(path)
        if row_count:
            break
    return (
        unique_join(file_names, separator="; "),
        dataset_file_types(selected_files),
        header_text,
        column_count,
        row_count,
        guess_target_column(header_text),
    )


def business_problem(title: str, description: str, tags: str, search_text: str) -> str:
    evidence = f"{title} {description} {tags} {search_text}".lower()
    if "churn" in evidence:
        return "Customer churn analysis / retention modeling"
    if "fraud" in evidence:
        return "Fraud detection / financial risk analysis"
    if "loan" in evidence or "credit" in evidence:
        return "Credit risk / loan decision analysis"
    if "bank" in evidence or "banking" in evidence:
        return "Banking customer behavior analysis"
    hits = sorted(word for word in BUSINESS_WORDS if word in evidence)
    return f"Business/data analysis relevance: {', '.join(hits[:8])}" if hits else ""


def project_fit_score(row: dict[str, Any], search_text: str) -> int:
    evidence = " ".join(str(value) for value in row.values()) + " " + search_text
    lower = evidence.lower()

    def has_real_value(column: str) -> bool:
        value = clean_text(row.get(column, ""))
        return bool(value and not value.startswith("Not provided"))

    score = 3
    if has_real_value("column_headers"):
        score += 2
    if has_real_value("visible_downloads") or has_real_value("visible_votes") or has_real_value("api_downloads"):
        score += 1
    if has_real_value("license"):
        score += 1
    if has_real_value("usability"):
        score += 1
    if any(word in lower for word in BUSINESS_WORDS):
        score += 2
    if "toy dataset" in lower or "beginner" in lower:
        score -= 1
    return max(1, min(10, score))


def recommendation(row: dict[str, Any], skipped_download_reason: str = "") -> str:
    if skipped_download_reason:
        return skipped_download_reason

    notes = [f"Fit score {row.get('project_fit_score', '')}/10."]
    if row.get("business_problem"):
        notes.append(row["business_problem"])
    if row.get("column_headers"):
        notes.append("Real file column headers extracted.")
    else:
        notes.append("No readable data-file headers found.")
    if row.get("target_column_guess"):
        notes.append(f"Likely target column: {row['target_column_guess']}.")
    if row.get("visible_downloads"):
        notes.append(f"Visible downloads: {row['visible_downloads']}.")
    return " ".join(notes)


def authenticate_kaggle() -> Any:
    from kaggle.api.kaggle_api_extended import KaggleApi

    api = KaggleApi()
    api.authenticate()
    return api


def dataset_details(api: Any, ref: str) -> tuple[Any, str]:
    if not api or not ref:
        return None, "Kaggle API unavailable or dataset reference missing."
    try:
        return api.dataset_view(ref), ""
    except AttributeError:
        # Some Kaggle package versions do not expose dataset_view. This is fine:
        # visible search-card data plus downloaded file headers still provide the
        # important selector fields.
        return None, ""
    except Exception as exc:
        return None, f"Detail metadata unavailable: {type(exc).__name__}: {exc}"


def download_dataset(api: Any, ref: str, target_dir: Path) -> str:
    if not api:
        return "Download skipped because Kaggle API is unavailable."
    try:
        target_dir.mkdir(parents=True, exist_ok=True)
        api.dataset_download_files(ref, path=str(target_dir), unzip=True, quiet=True)
        return ""
    except Exception as exc:
        return f"Download failed: {type(exc).__name__}: {exc}"


def api_search_fallback(api: Any, search_text: str, max_rows: int) -> list[dict[str, Any]]:
    if not api:
        return []
    try:
        datasets = api.dataset_list(search=search_text)
    except Exception:
        return []
    rows = []
    for index, dataset in enumerate(list(datasets or [])[:max_rows], start=1):
        ref = clean_text(get_attr(dataset, ["ref", "id"]))
        if not ref:
            owner = clean_text(get_attr(dataset, ["ownerName", "owner", "creatorName"]))
            slug = clean_text(get_attr(dataset, ["datasetSlug", "slug"]))
            ref = f"{owner}/{slug}" if owner and slug else ""
        if ref:
            rows.append({
                "search_rank": index,
                "dataset_title": clean_text(get_attr(dataset, ["title", "subtitle"])) or ref,
                "dataset_ref": ref,
                "result_url": f"{KAGGLE_BASE_URL}/datasets/{ref}",
                "result_type": "Dataset",
            })
    return rows


def row_from_visible_card(api: Any, card: dict[str, Any], search_text: str, no_download: bool) -> dict[str, Any]:
    ref = card.get("dataset_ref", "")
    detail, detail_note = dataset_details(api, ref)
    safe_ref_name = re.sub(r"[^A-Za-z0-9_.-]+", "_", ref or f"rank_{card.get('search_rank', '')}")
    dataset_dir = DOWNLOAD_DIR / safe_ref_name
    row: dict[str, Any] = {column: "" for column in COLUMNS}
    row.update({key: card.get(key, "") for key in row.keys() if key in card})

    sources = [detail] if detail else []
    row["api_author"] = clean_text(first_attr(sources, ["ownerName", "creatorName", "owner"]))
    row["updated_date"] = not_available(first_attr(sources, ["lastUpdated", "lastUpdatedDate", "updated"]))
    row["api_votes"] = not_available(safe_number(first_attr(sources, ["voteCount", "votes"])))
    row["api_downloads"] = not_available(safe_number(first_attr(sources, ["downloadCount", "downloads"])))
    row["api_views"] = not_available(safe_number(first_attr(sources, ["viewCount", "views"])))
    row["usability"] = (
        card.get("_visible_usability")
        or not_available(safe_number(first_attr(sources, ["usabilityRating", "usability"])))
    )
    row["license"] = not_available(first_attr(sources, ["licenseName", "license"]))
    row["update_frequency"] = not_available(first_attr(sources, ["expectedUpdateFrequency", "updateFrequency", "frequency"]))
    row["about_dataset"] = not_available(first_attr(sources, ["description", "subtitle"], row.get("visible_summary", "")))
    row["tags"] = tag_names(first_attr(sources, ["tags"], []) or [])

    size_bytes = safe_size_bytes(detail)
    row["size_mb"] = format_size_mb(size_bytes) or card.get("_visible_size_mb") or "Not provided by Kaggle API"
    skipped_download_reason = ""
    if no_download:
        skipped_download_reason = "Dataset download skipped by --no-download."
    elif size_bytes and size_bytes > MAX_DATASET_BYTES:
        skipped_download_reason = f"Skipped download because dataset is larger than {MAX_DATASET_BYTES // (1024 * 1024)} MB."
    elif ref and row.get("result_type") == "Dataset":
        skipped_download_reason = download_dataset(api, ref, dataset_dir)

    files = data_files(dataset_dir) if dataset_dir.exists() else []
    row["file_count"] = str(len(files)) if files else card.get("_visible_file_count", "")
    (
        row["file_name"],
        row["file_types"],
        row["column_headers"],
        row["column_count"],
        row["row_count_sample"],
        row["target_column_guess"],
    ) = extract_headers(files)
    if not row["file_types"]:
        row["file_types"] = card.get("_visible_file_types", "")
    row["business_problem"] = business_problem(
        row["dataset_title"],
        f"{row['about_dataset']} {row['visible_summary']}",
        row["tags"],
        search_text,
    )
    row["project_fit_score"] = project_fit_score(row, search_text)
    row["recommendation"] = recommendation(row, skipped_download_reason)

    notes = []
    if detail_note:
        notes.append(detail_note)
    if skipped_download_reason:
        notes.append(skipped_download_reason)
    if not row["column_headers"]:
        notes.append("No readable CSV/Excel/JSON/Parquet headers found.")
    row["metadata_notes"] = " ".join(notes)
    return row


def click_cookie_popup(page: Any) -> None:
    candidates = [
        page.get_by_role("button", name=COOKIE_BUTTON_PATTERN),
        page.locator("button").filter(has_text=COOKIE_BUTTON_PATTERN),
        page.locator('[role="button"]').filter(has_text=COOKIE_BUTTON_PATTERN),
    ]
    for buttons in candidates:
        try:
            count = min(buttons.count(), 5)
        except Exception:
            count = 0
        for index in range(count):
            try:
                button = buttons.nth(index)
                if button.is_visible(timeout=1000):
                    button.click(timeout=3000)
                    page.wait_for_timeout(3000)
                    log("Cookie popup handled")
                    return
            except Exception:
                continue


def scrape_visible_search_cards(search_url: str, max_rows: int, include_notebooks: bool) -> tuple[list[dict[str, Any]], str]:
    try:
        from playwright.sync_api import sync_playwright
    except Exception as exc:
        return [], f"Playwright unavailable: {type(exc).__name__}: {exc}"

    try:
        with sync_playwright() as playwright:
            browser = playwright.chromium.launch(headless=True)
            context = browser.new_context(viewport={"width": 1600, "height": 1200})
            page = context.new_page()
            page.set_default_timeout(30_000)
            log(f"Opening visible Kaggle search page: {search_url}")
            page.goto(search_url, wait_until="domcontentloaded")
            click_cookie_popup(page)
            for _ in range(8):
                page.mouse.wheel(0, 900)
                page.wait_for_timeout(500)

            raw_cards = page.locator("a[href]").evaluate_all(
                """
                anchors => {
                  const wanted = [];
                  const seen = new Set();
                  for (const a of anchors) {
                    const href = a.href || "";
                    const path = new URL(href, location.href).pathname;
                    const isDataset = path.startsWith("/datasets/");
                    const isNotebook = path.startsWith("/code/");
                    if (!isDataset && !isNotebook) continue;
                    const parts = path.split("/").filter(Boolean);
                    if (parts.length < 3) continue;
                    const normalized = `${location.origin}/${parts.slice(0, 3).join("/")}`;
                    if (seen.has(normalized)) continue;
                    seen.add(normalized);

                    const best = a.closest("li") || a.closest('[role="listitem"]') || a.parentElement || a;
                    wanted.push({
                      href: normalized,
                      anchorText: (a.innerText || a.textContent || "").trim(),
                      text: (best.innerText || best.textContent || "").trim()
                    });
                  }
                  return wanted;
                }
                """
            )
            context.close()
            browser.close()
    except Exception as exc:
        return [], f"Visible scrape failed: {type(exc).__name__}: {exc}"

    cards: list[dict[str, Any]] = []
    seen: set[str] = set()
    for raw in raw_cards:
        card_type = result_type_from_url(clean_text(raw.get("href")))
        if card_type != "Dataset" and not include_notebooks:
            continue
        card = parse_card(raw, len(cards) + 1)
        if not card["dataset_ref"] or card["result_url"] in seen:
            continue
        seen.add(card["result_url"])
        cards.append(card)
        if len(cards) >= max_rows:
            break
    return cards, ""


def merge_card_data(existing: dict[str, Any], incoming: dict[str, Any]) -> dict[str, Any]:
    """Merge global-search and dataset-page versions of the same card."""
    merged = dict(existing)
    for key, value in incoming.items():
        if value not in ("", None) and not merged.get(key):
            merged[key] = value
    # Prefer dataset-page title if the global anchor accidentally includes extra text.
    if incoming.get("dataset_title") and len(str(incoming["dataset_title"])) < len(str(merged.get("dataset_title", ""))):
        merged["dataset_title"] = incoming["dataset_title"]
    return merged


def collect_visible_cards(search_url: str, search_text: str, max_rows: int, include_notebooks: bool) -> tuple[list[dict[str, Any]], str]:
    """Scrape original search plus dataset-only page to get more than one dataset result."""
    urls = [search_url]
    if not include_notebooks:
        ds_url = dataset_search_url(search_text)
        if ds_url not in urls:
            urls.append(ds_url)

    by_url: dict[str, dict[str, Any]] = {}
    notes: list[str] = []
    for url in urls:
        cards, note = scrape_visible_search_cards(url, max_rows, include_notebooks)
        if note:
            notes.append(f"{url}: {note}")
        for card in cards:
            key = card.get("result_url", "")
            if not key:
                continue
            if key in by_url:
                by_url[key] = merge_card_data(by_url[key], card)
            else:
                by_url[key] = card
            if len(by_url) >= max_rows:
                break
        if len(by_url) >= max_rows:
            break

    cards = list(by_url.values())[:max_rows]
    for index, card in enumerate(cards, start=1):
        card["search_rank"] = index
    return cards, " ".join(notes)


def polish_excel(path: Path) -> None:
    workbook = load_workbook(path)
    sheet = workbook.active
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    wrap_columns = {"K", "Z", "AC", "AF", "AG"}
    preferred_widths = {
        "B": 32, "C": 28, "D": 45, "K": 55, "L": 32, "Z": 60,
        "AC": 45, "AF": 55, "AG": 45,
    }
    for column_cells in sheet.columns:
        column_letter = column_cells[0].column_letter
        max_length = max(len(str(cell.value or "")) for cell in column_cells[:200])
        sheet.column_dimensions[column_letter].width = preferred_widths.get(
            column_letter,
            min(max(max_length + 2, 12), 45),
        )
        for cell in column_cells[1:]:
            cell.alignment = Alignment(vertical="top", wrap_text=column_letter in wrap_columns)

    workbook.save(path)


def save_rows(rows: list[dict[str, Any]]) -> None:
    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    pd.DataFrame(rows, columns=COLUMNS).to_excel(OUTPUT_FILE, index=False, engine="openpyxl")
    polish_excel(OUTPUT_FILE)
    log(f"Saved Excel file: {OUTPUT_FILE}")


def error_row(message: str, search_text: str = "") -> dict[str, Any]:
    row: dict[str, Any] = {column: "" for column in COLUMNS}
    row["dataset_title"] = "No output generated"
    row["about_dataset"] = search_text
    row["project_fit_score"] = 1
    row["recommendation"] = "Fix the setup issue shown in metadata_notes, then rerun agent.py."
    row["metadata_notes"] = message
    return row


def main() -> None:
    args = parse_args()
    search_url, search_text = read_search_input(args)
    DOWNLOAD_DIR.mkdir(parents=True, exist_ok=True)
    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)

    log("Starting Hybrid Kaggle Dataset Selector")
    log(f"Search URL: {search_url}")
    log(f"Search text: {search_text}")

    cards, visible_note = collect_visible_cards(search_url, search_text, args.max, args.include_notebooks)
    log(f"Visible Kaggle cards found: {len(cards)}")

    api = None
    api_note = ""
    try:
        api = authenticate_kaggle()
        log("Kaggle API authenticated")
    except Exception as exc:
        api_note = f"Kaggle API unavailable: {type(exc).__name__}: {exc}"
        log(api_note)

    if not cards and api:
        log("Falling back to Kaggle API search results")
        cards = api_search_fallback(api, search_text, args.max)

    rows: list[dict[str, Any]] = []
    for index, card in enumerate(cards, start=1):
        log(f"[{index}/{len(cards)}] Processing {card.get('dataset_ref')}")
        row = row_from_visible_card(api, card, search_text, args.no_download)
        row["metadata_notes"] = " ".join(part for part in [visible_note, api_note, row.get("metadata_notes", "")] if part)
        rows.append(row)

    if not rows:
        rows = [error_row("No visible Kaggle dataset cards found. " + " ".join([visible_note, api_note]), search_text)]

    save_rows(rows)
    log(f"Finished. Saved {len(rows)} row(s) to {OUTPUT_FILE}")


def run() -> None:
    try:
        main()
    except Exception as exc:
        message = f"{type(exc).__name__}: {exc}"
        log(f"ERROR: {message}")
        try:
            args = parse_args()
            _, search_text = read_search_input(args)
        except Exception:
            search_text = ""
        save_rows([error_row(message, search_text)])


if __name__ == "__main__":
    run()
